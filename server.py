"""
MetaForge Document Service
- PDF rendering (WeasyPrint)
- Large file download + text extraction proxy (for files too large for Edge Functions)
"""
import os
import time
import io
import httpx

from fastapi import FastAPI, Request, Response, HTTPException
from pydantic import BaseModel
from typing import Optional

app = FastAPI(title="MetaForge Document Service", docs_url=None, redoc_url=None)

API_KEY = os.environ.get("WEASYPRINT_API_KEY", "")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
START_TIME = time.time()


def verify_api_key(request: Request):
    if not API_KEY:
        return
    key = request.headers.get("X-API-Key", "")
    if key != API_KEY:
        raise HTTPException(status_code=401, detail="Invalid API key")


@app.get("/health")
async def health():
    return {
        "status": "ok",
        "version": "1.0.0",
        "uptime_seconds": round(time.time() - START_TIME),
    }


@app.post("/render")
async def render(request: Request):
    verify_api_key(request)

    html_content = await request.body()
    if not html_content:
        raise HTTPException(status_code=400, detail="Empty HTML body")

    if len(html_content) > 50 * 1024 * 1024:  # 50MB limit
        raise HTTPException(status_code=413, detail="HTML body too large (max 50MB)")

    try:
        import weasyprint
        html = weasyprint.HTML(string=html_content.decode("utf-8"))
        pdf_bytes = html.write_pdf()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Render failed: {str(e)}")

    return Response(content=pdf_bytes, media_type="application/pdf")


async def select_page_with_claude(
    thumbnails: list[tuple[int, str]],  # [(page_num, base64_png), ...]
    api_key: str,
) -> int | None:
    """Use Claude Vision to pick the best architectural render page."""
    content = [
        {
            "type": "text",
            "text": (
                "Select the best page to use as a \"Site Plan\" image in a professional "
                "construction contract document.\n\n"
                "Pick a page showing:\n"
                "- A clean 3D architectural render of a house/villa exterior (photorealistic)\n"
                "- An aerial/bird's eye render showing the full property with landscaping\n"
                "- A perspective view of the villa with trees, driveway, garden\n"
                "- No dimension lines, no room labels, no measurement annotations\n\n"
                "Do NOT pick:\n"
                "- Floor plans with dimensions/measurements/room labels\n"
                "- Elevation drawings with annotations\n"
                "- BOQ tables, pricing tables, or text-heavy pages\n"
                "- Scope summary pages with tables/grids\n"
                "- Title/cover pages with just text or logos\n"
                "- Material/mood boards with multiple small images\n"
                "- Site photos (real photos, not renders)\n\n"
                "Below are the pages. Each is labeled with its page number.\n"
            ),
        }
    ]

    for page_num, b64 in thumbnails:
        content.append({"type": "text", "text": f"Page {page_num}:"})
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/png", "data": b64},
        })

    content.append({
        "type": "text",
        "text": (
            "\nRespond with ONLY the page number (0-indexed integer) of the best "
            "3D architectural render. If none of the pages contain a suitable "
            "render image, respond with -1.\n\nExample: 10"
        ),
    })

    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": "claude-sonnet-4-20250514",
                    "max_tokens": 32,
                    "temperature": 0,
                    "messages": [{"role": "user", "content": content}],
                },
            )
            resp.raise_for_status()

        data = resp.json()
        answer = data["content"][0]["text"].strip()
        page_num = int(answer)
        print(f"[Claude Vision] Selected page {page_num}")
        return page_num if page_num >= 0 else None
    except Exception as e:
        print(f"[Claude Vision] Failed: {e}")
        return None


def score_pages_heuristic(doc, max_pages: int, default_page: int = 0) -> int:
    """Fallback heuristic scorer — keyword + color analysis."""
    import gc
    from PIL import Image as PILImage

    PLAN_KEYWORDS = [
        "floor plan", "site plan", "ground floor", "first floor",
        "second floor", "roof plan", "layout plan", "proposed plan",
        "elevation", "landscape plan", "master plan",
    ]
    BOQ_KEYWORDS = [
        "bill of quantities", "boq description", "material type",
        "drawing reference", "number of items", "unit rate",
    ]

    best_score = -999999
    best_page = default_page

    for page_num in range(max_pages):
        try:
            page = doc[page_num]
            rect = page.rect
            scale = min(100 / rect.width, 100 / rect.height)
            pix = page.get_pixmap(matrix=__import__("fitz").Matrix(scale, scale), alpha=False)
            score = 0

            if rect.width > rect.height:
                score += 1000

            try:
                page_text = page.get_text("text").lower()[:2000]
            except Exception:
                page_text = ""

            plan_hits = sum(1 for kw in PLAN_KEYWORDS if kw in page_text)
            boq_hits = sum(1 for kw in BOQ_KEYWORDS if kw in page_text)
            if plan_hits > 0:
                score += 5000 + plan_hits * 500
            if boq_hits > 0:
                score -= 4000 + boq_hits * 300
            if max_pages > 4:
                if page_num == 0 and plan_hits == 0:
                    score -= 500
                if page_num >= max_pages - 1:
                    score -= 500

            thumb_bytes = pix.tobytes("ppm")
            pil_thumb = PILImage.open(io.BytesIO(thumb_bytes)).convert("RGB")
            colors = pil_thumb.getcolors(maxcolors=5000)
            if colors:
                score += min(len(colors), 1500)
                total = sum(c for c, _ in colors)
                white = sum(c for c, color in colors if sum(color) > 700)
                if total > 0 and white / total > 0.90:
                    score -= 2000

            if score > best_score:
                best_score = score
                best_page = page_num

            pix = None
            pil_thumb = None
            gc.collect()
        except Exception:
            continue

    return best_page


class ExtractImageRequest(BaseModel):
    url: str
    token: str
    filename: str
    page: Optional[int] = 0
    min_width: Optional[int] = 400
    min_height: Optional[int] = 300


@app.post("/extract-image")
async def extract_image(req: ExtractImageRequest, request: Request):
    """Download a PDF and extract the best architectural render as site plan image."""
    verify_api_key(request)

    try:
        # Download file
        async with httpx.AsyncClient(timeout=300.0) as client:
            resp = await client.get(
                req.url,
                headers={"Authorization": f"Bearer {req.token}"},
                follow_redirects=True,
            )
            resp.raise_for_status()
            file_bytes = resp.content

        ext = req.filename.rsplit(".", 1)[-1].lower() if "." in req.filename else ""

        if ext != "pdf":
            raise HTTPException(status_code=400, detail="Only PDF files supported for image extraction")

        import fitz
        import base64
        import gc
        from PIL import Image as PILImage

        doc = fitz.open(stream=file_bytes, filetype="pdf")
        max_pages = min(doc.page_count, 15)

        # Render all pages as 300px-wide thumbnails for Claude Vision
        thumbnails: list[tuple[int, str]] = []
        for page_num in range(max_pages):
            try:
                page = doc[page_num]
                rect = page.rect
                scale = min(300 / rect.width, 300 / rect.height)
                pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
                thumb_png = pix.tobytes("png")
                thumb_b64 = base64.b64encode(thumb_png).decode("ascii")
                thumbnails.append((page_num, thumb_b64))
                pix = None
            except Exception:
                continue
        gc.collect()

        # Select best page — Claude Vision with heuristic fallback
        best_page_num = req.page

        if ANTHROPIC_API_KEY and len(thumbnails) > 1:
            claude_pick = await select_page_with_claude(thumbnails, ANTHROPIC_API_KEY)
            if claude_pick is not None and 0 <= claude_pick < max_pages:
                best_page_num = claude_pick
                print(f"[extract-image] Claude Vision selected page {best_page_num}")
            else:
                best_page_num = score_pages_heuristic(doc, max_pages, req.page)
                print(f"[extract-image] Heuristic fallback selected page {best_page_num}")
        else:
            best_page_num = score_pages_heuristic(doc, max_pages, req.page)
            print(f"[extract-image] Heuristic selected page {best_page_num} (no API key)")

        # Render selected page at good quality (1.5x)
        try:
            page = doc[best_page_num]
            mat = fitz.Matrix(1.5, 1.5)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            best_image = pix.tobytes("png")
            best_page = best_page_num
            pix = None
        except Exception:
            page = doc[0]
            pix = page.get_pixmap(matrix=fitz.Matrix(1, 1), alpha=False)
            best_image = pix.tobytes("png")
            best_page = 0

        doc.close()
        gc.collect()

        # Resize if too large (max 1200px wide)
        from PIL import Image
        img = Image.open(io.BytesIO(best_image))
        max_width = 1200
        if img.width > max_width:
            ratio = max_width / img.width
            img = img.resize((max_width, int(img.height * ratio)), Image.LANCZOS)

        buf = io.BytesIO()
        if img.mode == "RGBA":
            img = img.convert("RGB")
        img.save(buf, format="JPEG", quality=85)
        optimized = buf.getvalue()

        b64 = base64.b64encode(optimized).decode("utf-8")

        return {
            "filename": req.filename,
            "page": best_page,
            "image_size": len(optimized),
            "base64": b64,
            "format": "jpeg",
        }

    except httpx.HTTPError as e:
        raise HTTPException(status_code=502, detail=f"Download failed: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Image extraction failed: {str(e)}")


class ExtractRequest(BaseModel):
    url: str
    token: str
    filename: str
    max_text_chars: Optional[int] = 50000


@app.post("/extract-text")
async def extract_text(req: ExtractRequest, request: Request):
    """Download a file from URL and extract text content.
    Works for PDFs of ANY size — no timeout limits on Render."""
    verify_api_key(request)

    try:
        # Download file from Slack (no timeout limit on Render)
        async with httpx.AsyncClient(timeout=300.0) as client:
            resp = await client.get(
                req.url,
                headers={"Authorization": f"Bearer {req.token}"},
                follow_redirects=True,
            )
            resp.raise_for_status()
            file_bytes = resp.content

        ext = req.filename.rsplit(".", 1)[-1].lower() if "." in req.filename else ""
        size_mb = len(file_bytes) / 1024 / 1024
        extracted_text = ""

        if ext == "pdf":
            try:
                import fitz  # PyMuPDF
                doc = fitz.open(stream=file_bytes, filetype="pdf")
                pages_text = []
                for page_num in range(min(doc.page_count, 100)):
                    page = doc[page_num]
                    pages_text.append(f"--- Page {page_num + 1} ---\n{page.get_text()}")
                extracted_text = "\n".join(pages_text)

                # OCR fallback: if text extraction got almost nothing, the PDF is
                # scanned/image-based (like Tatiana's contractor quotes). Use Claude
                # Vision to read the pages. Cheap (~$0.03) and catches all the prices
                # and descriptions that text extraction misses.
                clean_text = extracted_text.replace("---", "").replace("Page", "").strip()
                if len(clean_text) < 200 and ANTHROPIC_API_KEY and doc.page_count > 0:
                    print(f"[extract-text] Scanned PDF detected ({len(clean_text)} chars). Using Claude Vision OCR...")
                    import base64
                    import gc
                    content = [{"type": "text", "text": (
                        "This is a scanned construction document. Extract ALL text exactly as printed. "
                        "For each page, output the text preserving the structure (headers, line items, "
                        "amounts, descriptions). Use ONLY printed text — ignore handwritten annotations, "
                        "pencil marks, and margin notes. Preserve exact AED amounts as printed."
                    )}]
                    max_ocr_pages = min(doc.page_count, 10)
                    for pg in range(max_ocr_pages):
                        page = doc[pg]
                        scale = min(1200 / page.rect.width, 1200 / page.rect.height)
                        pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
                        png_b64 = base64.b64encode(pix.tobytes("png")).decode("ascii")
                        content.append({"type": "text", "text": f"--- Page {pg + 1} ---"})
                        content.append({"type": "image", "source": {"type": "base64", "media_type": "image/png", "data": png_b64}})
                        pix = None
                    gc.collect()

                    try:
                        async with httpx.AsyncClient(timeout=60.0) as ocr_client:
                            ocr_resp = await ocr_client.post(
                                "https://api.anthropic.com/v1/messages",
                                headers={
                                    "x-api-key": ANTHROPIC_API_KEY,
                                    "anthropic-version": "2023-06-01",
                                    "content-type": "application/json",
                                },
                                json={
                                    "model": "claude-sonnet-4-20250514",
                                    "max_tokens": 8192,
                                    "temperature": 0,
                                    "messages": [{"role": "user", "content": content}],
                                },
                            )
                            ocr_data = ocr_resp.json()
                            extracted_text = "[OCR via Claude Vision]\n" + ocr_data["content"][0]["text"]
                            print(f"[extract-text] OCR extracted {len(extracted_text)} chars")
                    except Exception as ocr_err:
                        print(f"[extract-text] OCR failed: {ocr_err}, using minimal text extraction")

                doc.close()
            except Exception:
                try:
                    import pdfplumber
                    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                        pages_text = []
                        for i, page in enumerate(pdf.pages[:100]):
                            text = page.extract_text() or ""
                            pages_text.append(f"--- Page {i + 1} ---\n{text}")
                        extracted_text = "\n".join(pages_text)
                except Exception as e2:
                    extracted_text = f"[PDF text extraction failed: {str(e2)}]"

        elif ext in ("xlsx", "xls", "csv"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                sheets_text = []
                for name in wb.sheetnames:
                    ws = wb[name]
                    sheets_text.append(f"=== Sheet: {name} ===")
                    for row in ws.iter_rows(values_only=True):
                        vals = [str(v) if v is not None else "" for v in row]
                        if any(v.strip() for v in vals):
                            sheets_text.append(" | ".join(vals))
                extracted_text = "\n".join(sheets_text)
            except Exception as e:
                extracted_text = f"[Excel extraction failed: {str(e)}]"

        else:
            extracted_text = f"[Binary file: {req.filename} ({size_mb:.1f}MB, .{ext})]"

        # Truncate if needed
        if len(extracted_text) > req.max_text_chars:
            extracted_text = extracted_text[:req.max_text_chars] + "\n\n[... truncated ...]"

        return {
            "filename": req.filename,
            "size_mb": round(size_mb, 1),
            "extension": ext,
            "text_length": len(extracted_text),
            "text": extracted_text,
        }

    except httpx.HTTPError as e:
        raise HTTPException(status_code=502, detail=f"Download failed: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Extraction failed: {str(e)}")
